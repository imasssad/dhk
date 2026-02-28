from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
import io
import xlsxwriter
import openpyxl

class POPriceUpdateWizard(models.TransientModel):
    _name = 'po.price.update.wizard'
    _description = 'PO Price Update Wizard'
    
    sale_order_id = fields.Many2one('sale.order', string='Sales Order', required=True)
    action_type = fields.Selection([
        ('export', 'Export'),
        ('import', 'Import')
    ], string='Action', required=True, default='export')
    
    # For export
    excel_file = fields.Binary(string='Excel File', readonly=True)
    file_name = fields.Char(string='File Name', readonly=True)
    
    # For import
    import_file = fields.Binary(string='Upload Excel File')
    import_file_name = fields.Char(string='Import File Name')
    
    update_summary = fields.Text(string='Update Summary', readonly=True)
    
    def action_export_po(self):
        """Export related PO lines to Excel"""
        self.ensure_one()
        
        # Get all purchase order lines related to this SO
        # Use sudo() to bypass custom access rules that might conflict
        po_lines = self.env['purchase.order.line'].sudo().search([
            ('sale_line_id', 'in', self.sale_order_id.order_line.ids)
        ])
        
        if not po_lines:
            raise UserError('No related Purchase Orders found for this Sales Order.')
        
        # Create Excel file
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('PO Lines')
        
        # Header format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D3D3D3',
            'border': 1
        })
        
        # Define columns
        headers = [
            'PO Line ID',
            'PO Number',
            'PO State',
            'SO Number',
            'Vendor',
            'Product Code',
            'Product Name',
            'Original Unit Price',
            'Updated Unit Price'
        ]
        
        # Write headers
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Write data
        row = 1
        for line in po_lines:
            worksheet.write(row, 0, line.id)
            worksheet.write(row, 1, line.order_id.name)
            worksheet.write(row, 2, line.order_id.state)
            worksheet.write(row, 3, line.sale_line_id.order_id.name if line.sale_line_id else '')
            worksheet.write(row, 4, line.order_id.partner_id.name)
            worksheet.write(row, 5, line.product_id.default_code or '')
            worksheet.write(row, 6, line.product_id.name)
            worksheet.write(row, 7, line.price_unit)
            worksheet.write(row, 8, line.price_unit)  # To be updated
            row += 1
        
        # Adjust column widths
        worksheet.set_column('A:A', 12)  # PO Line ID
        worksheet.set_column('B:C', 20)  # PO Number, PO State
        worksheet.set_column('D:D', 20)  # SO Number
        worksheet.set_column('E:E', 25)  # Vendor
        worksheet.set_column('F:G', 30)  # Product Code, Product Name
        worksheet.set_column('H:I', 18)  # Original Price, Updated Price
        
        workbook.close()
        
        # Save file
        excel_data = output.getvalue()
        file_name = f'PO_Export_{self.sale_order_id.name.replace("/", "_")}.xlsx'
        
        self.write({
            'excel_file': base64.b64encode(excel_data),
            'file_name': file_name,
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'po.price.update.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
    
    def action_import_po_prices(self):
        """Import updated prices from Excel and update PO and SO"""
        self.ensure_one()
        
        if not self.import_file:
            raise UserError('Please upload an Excel file to import.')
        
        # Decode the file
        file_data = base64.b64decode(self.import_file)
        workbook = openpyxl.load_workbook(io.BytesIO(file_data))
        worksheet = workbook.active
        
        updated_lines = []
        errors = []
        
        # Skip header row
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:  # Skip empty rows
                    continue
                    
                po_line_id = int(row[0])  # PO Line ID column
                new_price = float(row[8])  # Updated Unit Price column
                
                # Use sudo() to bypass custom access rules
                po_line = self.env['purchase.order.line'].sudo().browse(po_line_id)
                
                if not po_line.exists():
                    errors.append(f'Row {row_idx}: PO Line ID {po_line_id} not found.')
                    continue
                
                old_price = po_line.price_unit
                
                # Check if PO can be modified
                if po_line.order_id.state not in ['draft', 'sent', 'to approve']:
                    try:
                        po_line.order_id.button_draft()
                    except Exception as e:
                        errors.append(f'Row {row_idx}: Cannot reset PO {po_line.order_id.name} to draft. Error: {str(e)}')
                        continue
                
                # Update PO line price
                po_line.write({'price_unit': new_price})
                
                # Update SO line cost (purchase_price)
                if po_line.sale_line_id:
                    po_line.sale_line_id.sudo().write({'purchase_price': new_price})
                    updated_lines.append({
                        'po': po_line.order_id.name,
                        'product': po_line.product_id.name,
                        'old_price': old_price,
                        'new_price': new_price,
                        'so': po_line.sale_line_id.order_id.name
                    })
                
            except ValueError as e:
                errors.append(f'Row {row_idx}: Invalid data format. {str(e)}')
            except Exception as e:
                errors.append(f'Row {row_idx}: {str(e)}')
        
        # Create summary
        summary = f"Successfully updated {len(updated_lines)} PO lines.\n\n"
        
        if updated_lines:
            summary += "Details:\n"
            for line in updated_lines:
                summary += f"- PO: {line['po']} | Product: {line['product']}\n"
                summary += f"  Price: {line['old_price']} → {line['new_price']}\n"
                summary += f"  Updated in SO: {line['so']}\n\n"
        
        if errors:
            summary += f"\n\nErrors ({len(errors)}):\n"
            summary += "\n".join(errors)
        
        self.update_summary = summary
        
        # Refresh SO to show updated costs
        self.sale_order_id.invalidate_recordset()
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'po.price.update.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
    
    def action_download_template(self):
        """Download the exported file"""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/po.price.update.wizard/{self.id}/excel_file/{self.file_name}?download=true',
            'target': 'self',
        }
